
import logging
from django.http import Http404, JsonResponse
from openpyxl import load_workbook
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import generics,status,viewsets
from django.db.models import Q

from .models import Reconciliation, ReconciliationLog, Transactions, UploadedFile
from .serializers import ReconcileSerializer, UploadedFileSerializer

from .utils import (select_exceptions,select_reversals,date_range, backup_refs, db_transactions, pre_processing,bank_reconciliation, recon_logs_req, unserializable_floats, use_cols, update_reconciliation,update_exception_flag, insert_recon_stats)
import pandas as pd
import os

# Create your views here.

class ReconcileView(APIView):
    serializer_class = ReconcileSerializer
    # permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        
        if serializer.is_valid():
            uploaded_file = serializer.validated_data['file']
            swift_code = serializer.validated_data['swift_code']

            # Save the uploaded file temporarily
            temp_file_path = "temp_file.xlsx"
            with open(temp_file_path, "wb") as buffer:
                buffer.write(uploaded_file.read())

            try:
                # Call the main function with the path of the saved file and the swift code
                merged_df, reconciled_data, succunreconciled_data, exceptions, feedback, requestedRows, UploadedRows, date_range_str = reconcileMain(
                    temp_file_path, swift_code)
                
                # Perform clean up: remove the temporary file after processing
                os.remove(temp_file_path)
                
                data = {
                    "reconciledRows": len(reconciled_data),
                    "unreconciledRows": len(succunreconciled_data),
                    "exceptionsRows": len(exceptions),
                    "feedback": feedback,
                    "RequestedRows": requestedRows,
                    "UploadedRows": UploadedRows,
                    "min_max_DateRange": date_range_str
                }

                return Response(data, status=status.HTTP_200_OK)

            except Exception as e:
                # If there's an error during the process, ensure the temp file is removed
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                
                # Return error as response
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
# class ReconcileView(APIView):
#     serializer_class = ReconcileSerializer
#     # permission_classes = [IsAuthenticated]

#     def post(self, request):
#         serializer = self.serializer_class(data=request.data)
#         if serializer.is_valid():
#             uploaded_file = serializer.validated_data['file']
#             swift_code = serializer.validated_data['swift_code']

#             # Save the uploaded file temporarily
#             temp_file_path = "temp_file.xlsx"
            
#             with open(temp_file_path, "wb") as buffer:
#                 buffer.write(uploaded_file.read())
               

#             try:
#                 # Call the main function with the path of the saved file and the swift code
                
#                 result = reconcile_main_view(temp_file_path,swift_code)
#                 print(result)
#                 return result
                
#                 # return Response(result, status=status.HTTP_200_OK)

#             except Exception as e:
#                 # If there's an error during the process, ensure the temp file is removed
#                 if os.path.exists(temp_file_path):
#                     os.remove(temp_file_path)
                
#                 # Return error as response
#                 return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class UploadedFilesViewset(viewsets.ModelViewSet):
    queryset = UploadedFile.objects.all()
    serializer_class = UploadedFileSerializer
    def create(self, request, *args, **kwargs):
        user = request.user
        file = request.FILES['file']
        wb = load_workbook(file)
        sheet = wb.get_sheet_by_name("Sheet1")
        start_row = 2
        count = 0
        for _ in sheet.iter_rows(min_row=start_row,max_row=10):
            row_no = start_row+count
            time = sheet.cell(row_no,1).value
            transaction_type = sheet.cell(row_no,2).value
            amount = sheet.cell(row_no,3).value
            abc_reference = sheet.cell(row_no,4).value
            recon = Reconciliation(
                date_time=time,
                last_modified_by_user=user,
                trn_ref=abc_reference,
                )
            recon.save()
            print(time,transaction_type,amount,abc_reference)
            count+=1
        
        return super().create(request, *args, **kwargs) 


         
def reconcileMain(path, swift_code):
    datadump = pd.DataFrame()  # Initializing datadump as an empty DataFrame at the start

    try:
        # Ensure the file exists and is loaded correctly
        if not os.path.exists(path):
            logging.error("File does not exist.")
            return None, None, None, None, None, None, None, None
        
        # Read the uploaded dataset from Excel
        uploaded_df = pd.read_excel(path, usecols=[0, 1, 2, 3], skiprows=0)

        # Check if uploaded_df is loaded correctly
        if uploaded_df is None or uploaded_df.empty:
            logging.error("uploaded_df is empty or None.")
            return None, None, None, None, None, None, None, None

        # Format the 'Date' column
        min_date, max_date = date_range(uploaded_df, 'Date')
        date_range_str = f"{min_date},{max_date}"

        uploaded_df = backup_refs(uploaded_df, 'ABC Reference')
        uploaded_df['Response_code'] = '0'
        UploadedRows = len(uploaded_df)
        
        # Clean and format columns in the uploaded dataset
        uploaded_df_processed = pre_processing(uploaded_df)

        # Retrieve data using Django ORM and the Transactions model
        datadump = Transactions.objects.filter(
            Q(ISSUER_CODE=swift_code) | Q(ACQUIRER_CODE=swift_code),
            DATE_TIME__range=[min_date, max_date]
        ).exclude(
            REQUEST_TYPE__in=['1420', '1421'],
            TXN_TYPE__in=['ACI', 'AGENTFLOATINQ', 'BI', 'MINI'],
            AMOUNT=0
        ).distinct().values('DATE_TIME', 'BATCH', 'TRN_REF', 'TXN_TYPE', 'ISSUER_CODE', 'ACQUIRER_CODE', 'AMOUNT', 'RESPONSE_CODE')
        datadump = pd.DataFrame.from_records(datadump)

        if datadump is not None:
            datadump = backup_refs(datadump, 'TRN_REF')
            requestedRows = len(datadump[datadump['RESPONSE_CODE'] == '0'])

        # Clean and format columns in the datadump
        db_preprocessed = pre_processing(datadump)
        
        merged_df, reconciled_data, succunreconciled_data, exceptions = bank_reconciliation(uploaded_df_processed, db_preprocessed)
     
        succunreconciled_data = use_cols(succunreconciled_data)
        reconciled_data = use_cols(reconciled_data)

        feedback = update_reconciliation(reconciled_data, swift_code)
        
        exceptions_feedback = None
        # Check if exceptions DataFrame is not empty, if not empty then update exception flag
        if not exceptions.empty:
            exceptions_feedback = update_exception_flag(exceptions, swift_code)
        else:
            exceptions_feedback = "No exceptions to update."
            
        insert_recon_stats(
            swift_code, len(reconciled_data), len(succunreconciled_data),
            len(exceptions), feedback, (requestedRows), (UploadedRows),
            date_range_str
        )
        
        return merged_df, reconciled_data, succunreconciled_data, exceptions, feedback, requestedRows, UploadedRows, date_range_str

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        return JsonResponse({"error": "No data dump available."}, status=404)

# def reconcile_main_view (path, Swift_code_up):
#     try:
#         global reconciled_data, succunreconciled_data  # Indicate these are global variables
        
#         uploaded_df = pd.read_excel(path , usecols=[0, 1, 2, 3], skiprows=0)

#         # Now, you can use strftime to format the 'Date' column
#         min_date, max_date = date_range(uploaded_df, 'Date')

#         date_range_str = f"{min_date},{max_date}"

#         uploaded_df = backup_refs(uploaded_df, 'ABC Reference')
#         uploaded_df['Response_code'] = '0'
#         UploadedRows = len(uploaded_df)
        
#         # Clean and format columns in the uploaded dataset
#         uploaded_df_processed = pre_processing(uploaded_df)
        
        
#     #     query = f"""
#     #      SELECT DISTINCT DATE_TIME, BATCH,TRN_REF, TXN_TYPE, ISSUER_CODE, ACQUIRER_CODE,
#     #             AMOUNT, RESPONSE_CODE
#     #      FROM Transactions
#     #      WHERE (ISSUER_CODE = '{Swift_code_up}' OR ACQUIRER_CODE = '{Swift_code_up}')
#     #          AND CONVERT(DATE, DATE_TIME) BETWEEN '{min_date}' AND '{max_date}'
#     #          AND REQUEST_TYPE NOT IN ('1420','1421')
#     #         AND AMOUNT <> 0
#     #         AND TXN_TYPE NOT IN ('ACI','AGENTFLOATINQ','BI','MINI')
#     #  """


#     #     # Execute the SQL query
#     #     datadump = execute_query(server, database, username, password, query, query_type="SELECT")
#         datadump = db_transactions(min_date, max_date,Swift_code_up)
#         if datadump is not None:
#             datadump = backup_refs(datadump, 'TRN_REF')
#             requestedRows = len(datadump[datadump['RESPONSE_CODE'] == '0'])

#             # Clean and format columns in the datadump        
#             db_preprocessed = pre_processing(datadump)
                    
#             merged_df, reconciled_data, succunreconciled_data, exceptions = bank_reconciliation(uploaded_df_processed, db_preprocessed)  
#             succunreconciled_data = use_cols(succunreconciled_data) 
#             reconciled_data = use_cols(reconciled_data) 

#             feedback = update_reconciliation(reconciled_data, Swift_code_up)      
#             # Initialize exceptions_feedback with a default value
#             exceptions_feedback = None 
#             # Check if exceptions DataFrame is not empty, if not empty then update exception flag
#             # if not exceptions.empty:
#             # exceptions_feedback = update_exception_flag(exceptions)
#             # else:
#             #     exceptions_feedback = "No exceptions to update."
                
#             insert_recon_stats(
#                 Swift_code_up, Swift_code_up, len(reconciled_data), len(succunreconciled_data), 
#                 len(exceptions), feedback, (requestedRows), (UploadedRows), 
#                 date_range_str
#             ) 
            
#             print('Thank you, your reconciliation is complete. ' + feedback)

#             return merged_df, reconciled_data, succunreconciled_data, exceptions, feedback, requestedRows, UploadedRows, date_range_str

#     except Exception as e:
#         return JsonResponse({"error": "No data dump available."}, status=404)
        


    



























































# def ReconStats_req(request):
#     bank_id = request.GET.get('bank_id')  # Assuming you're passing bank_id as a GET parameter
#     if not bank_id:
#          return JsonResponse({'error': 'bank_id not provided'}, status=400)
    
#     logs_query = recon_logs_req(bank_id)
#     if logs_query is None:
#         return JsonResponse({'error': 'No records to display'}, status=500)
    
#     # Convert the QuerySet to a list of dictionaries
#     logs_list = list(logs_query.values())
    
#     return JsonResponse(logs_list, safe=False)

# class ReconStatsView(APIView):
    
#     def get(self, request, Swift_code_up):
            
#             try:
#                 # Assume recon_stats_req returns a list of dictionaries or None
#                 data = ReconStats_req(Swift_code_up)
#                 if data is None:
#                     return Response({'error': 'No data found'}, status=status.HTTP_404_NOT_FOUND)              
                
#                 return Response(data, status=status.HTTP_200_OK)

#             except Exception as e:
#                 return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
# def get_reversals(request):
#     bank_id = request.GET.get('bank_id')  # Assuming you're passing bank_id as a GET parameter
#     if not bank_id:
#         return JsonResponse({'error': 'bank_id not provided'}, status=400)
    
#     reversals_query = select_reversals(bank_id)
    
#     if reversals_query is None:
#         return JsonResponse({'error': 'No records to display'}, status=500)
    
#     # Convert the QuerySet to a list of dictionaries
#     reversals_list = list(reversals_query)
    
#     return JsonResponse(reversals_list, safe=False)

# class ReversalsView(APIView):
#     """
#     Retrieve reversal data.
#     """

#     def get(self, request, swift_code_up, *args, **kwargs):
#         # Use values from .env for database connection
#         data = get_reversals(swift_code_up)

#         # The data returned by select_reversals is assumed to be in a suitable format for JSON serialization

#         return Response(data, status=status.HTTP_200_OK)
    
# def get_exceptions(request):
#     bank_id = request.GET.get('bank_id')  # Assuming you're passing bank_id as a GET parameter
#     if not bank_id:
#         return JsonResponse({'error': 'bank_id not provided'}, status=400)
    
#     exceptions_query = select_exceptions(bank_id)
    
#     if exceptions_query is None:
#         return JsonResponse({'error': 'No records to display'}, status=500)
    
#     # Convert the QuerySet to a list of dictionaries and process RECON_STATUS
#     exceptions_list = []
#     for result in exceptions_query:
#         record = {
#             'DATE_TIME': result['DATE_TIME'],
#             'TRAN_DATE': result['TRAN_DATE'],
#             'TRN_REF': result['TRN_REF'],
#             'BATCH': result['BATCH'],
#             'ACQUIRER_CODE': result['ACQUIRER_CODE'],
#             'ISSUER_CODE': result['ISSUER_CODE'],
#             'EXCEP_FLAG': result['EXCEP_FLAG']
#         }
#         if result.get('ACQ_FLG') == 1 or result.get('ISS_FLG') == 1:
#             record['RECON_STATUS'] = 'Partly Reconciled'
#         elif result.get('ACQ_FLG') == 1 and result.get('ISS_FLG') == 1:
#             record['RECON_STATUS'] = 'Fully Reconciled'
#         exceptions_list.append(record)
    
#     return JsonResponse(exceptions_list, safe=False)

# class ExceptionsView(APIView):
#     """
#     Retrieve exceptions data.    """

#     def get(self, request, swift_code_up, *args, **kwargs):          

#         data = get_exceptions(swift_code_up)

#         # The data returned by select_exceptions is assumed to be in a suitable format for JSON serialization
#         return Response(data, status=status.HTTP_200_OK)

# class ReconciledDataView(APIView):
#     """
#     Retrieve reconciled data.
#     """

#     def get(self, request, *args, **kwargs):
#         global reconciled_data

#         if reconciled_data is not None:
#             reconciled_data_cleaned = unserializable_floats(reconciled_data)
#             data = reconciled_data_cleaned.to_dict(orient='records')
#             return Response(data, status=status.HTTP_200_OK)
#         else:
#             raise Http404("Reconciled data not found")

# class UnReconciledDataView(APIView):
    # """
    # Retrieve unreconciled data.
    # """

    # def get(self, request, *args, **kwargs):
    #     global succunreconciled_data

    #     if succunreconciled_data is not None:
            
    #         # reconciled_data_cleaned = unserializable_floats(reconciled_data)
    #         # data = reconciled_data_cleaned.to_dict(orient='records')

    #         unreconciled_data_cleaned = unserializable_floats(succunreconciled_data)
    #         data =  unreconciled_data_cleaned.to_dict(orient='records')
    #         return Response(data, status=status.HTTP_200_OK)
    #     else:
    #         raise Http404("Unreconciled data not found")  